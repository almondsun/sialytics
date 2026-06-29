"""Exportación atómica de un libro académico orientado al estudiante."""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .analytics import SemesterMetrics, calculate_papa, calculate_pappi, calculate_semester_metrics
from .models import AcademicRecord, OfficialValue
from .security import safe_spreadsheet_text, secure_output_permissions
from .validation import validate

GREEN = "1B5E20"
LIGHT_GREEN = "E8F5E9"
ORANGE = "F9A825"
BLUE = "1565C0"
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
HEADER_FONT = Font(color="FFFFFF", bold=True)
CARD_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)
INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")


def _display(value: OfficialValue) -> object:
    return safe_spreadsheet_text(value.display_value)


def _metric(value: float | None) -> float | None:
    return value


def _style_title(sheet: Worksheet, title: str, end_column: str) -> None:
    sheet.merge_cells(f"A1:{end_column}1")
    cell = sheet["A1"]
    cell.value = safe_spreadsheet_text(title)
    cell.fill = HEADER_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False


def _style_header(sheet: Worksheet, row: int, start: int, end: int) -> None:
    for cell in sheet.iter_cols(min_col=start, max_col=end, min_row=row, max_row=row):
        item = cell[0]
        item.fill = HEADER_FILL
        item.font = HEADER_FONT
        item.border = THIN_BORDER
        item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_range(sheet: Worksheet, min_row: int, max_row: int, max_col: int) -> None:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _safe_sheet_title(workbook: Workbook, title: str) -> str:
    base = INVALID_SHEET_TITLE.sub("_", title).strip(" '")[:31] or "Semestre"
    existing = {name.casefold() for name in workbook.sheetnames}
    if base.casefold() not in existing:
        return base
    for index in range(2, 1_000):
        suffix = f"-{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing:
            return candidate
    raise ValueError(f"No se pudo crear una hoja única para {title!r}")


def _official_indicators(record: AcademicRecord) -> list[tuple[str, object]]:
    return [
        (indicator.acronym or indicator.canonical_name, _display(indicator.official_value))
        for indicator in record.indicators
    ]


def _write_summary(
    workbook: Workbook, record: AcademicRecord, semesters: list[SemesterMetrics]
) -> Worksheet:
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Resumen"
    sheet.sheet_properties.tabColor = GREEN
    _style_title(sheet, record.student.program.name, "L")

    papa = calculate_papa(record.courses)
    pappi = calculate_pappi(record.courses)
    total_credits = sum(item.credits for item in semesters)
    official_indicators = _official_indicators(record)
    official_by_acronym = {
        label.casefold(): value
        for label, value in official_indicators
        if label.casefold() in {"papa", "pappi"}
    }
    other_official_indicators = [
        item for item in official_indicators if item[0].casefold() not in {"papa", "pappi"}
    ]
    cards: list[tuple[str, object]] = [
        ("PAPA", official_by_acronym.get("papa", _metric(papa))),
        ("PAPPI", official_by_acronym.get("pappi", _metric(pappi))),
        *other_official_indicators,
        ("Semestres", len(semesters)),
        ("Asignaturas", len(record.courses)),
        ("Créditos registrados", total_credits),
    ]
    for index, (label, value) in enumerate(cards):
        row = 3 + index // 4 * 2
        column = 1 + index % 4 * 3
        sheet.cell(row=row, column=column, value=cast(Any, safe_spreadsheet_text(label)))
        sheet.cell(row=row + 1, column=column, value=cast(Any, safe_spreadsheet_text(value)))
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        sheet.merge_cells(
            start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1
        )
        label_cell = sheet.cell(row=row, column=column)
        value_cell = sheet.cell(row=row + 1, column=column)
        label_cell.fill = HEADER_FILL
        label_cell.font = HEADER_FONT
        value_cell.fill = CARD_FILL
        value_cell.font = Font(bold=True, size=14)
        for cell in (label_cell, value_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    table_row = 9
    headers = ["Semestre", "PAPPI", "PAPA", "Créditos", "Asignaturas", "Aprobadas"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=table_row, column=column, value=header)
    _style_header(sheet, table_row, 1, len(headers))
    for semester in semesters:
        sheet.append(
            [
                semester.semester,
                _metric(semester.pappi),
                _metric(semester.cumulative_papa),
                semester.credits,
                semester.course_count,
                semester.passed,
            ]
        )
    if semesters:
        last_row = table_row + len(semesters)
        _style_data_range(sheet, table_row + 1, last_row, len(headers))

        average_chart = LineChart()
        average_chart.title = "Evolución de PAPPI y PAPA"
        average_chart.y_axis.title = "Calificación"
        average_chart.x_axis.title = "Semestre"
        average_chart.add_data(
            Reference(sheet, min_col=2, max_col=3, min_row=table_row, max_row=last_row),
            titles_from_data=True,
        )
        average_chart.set_categories(
            Reference(sheet, min_col=1, min_row=table_row + 1, max_row=last_row)
        )
        average_chart.height = 7
        average_chart.width = 14
        sheet.add_chart(average_chart, "H9")

        credits_chart = BarChart()
        credits_chart.title = "Créditos por semestre"
        credits_chart.y_axis.title = "Créditos"
        credits_chart.add_data(
            Reference(sheet, min_col=4, min_row=table_row, max_row=last_row),
            titles_from_data=True,
        )
        credits_chart.set_categories(
            Reference(sheet, min_col=1, min_row=table_row + 1, max_row=last_row)
        )
        credits_chart.height = 7
        credits_chart.width = 14
        sheet.add_chart(credits_chart, "H24")

    sheet.freeze_panes = "A10"
    for column in range(1, 13):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    return sheet


def _write_semester(workbook: Workbook, semester: SemesterMetrics) -> None:
    sheet = workbook.create_sheet(_safe_sheet_title(workbook, semester.semester))
    sheet.sheet_properties.tabColor = "43A047"
    _style_title(sheet, f"Semestre académico {semester.semester}", "P")

    cards: list[tuple[str, object]] = [
        ("PAPPI", _metric(semester.pappi)),
        ("PAPA al cierre", _metric(semester.cumulative_papa)),
        ("Créditos", semester.credits),
        ("Asignaturas", semester.course_count),
        ("Aprobadas", semester.passed),
        ("No aprobadas", semester.failed),
        ("Calificación más alta", _metric(semester.highest_grade)),
        ("Calificación mediana", _metric(semester.median_grade)),
    ]
    for index, (label, value) in enumerate(cards):
        column = 1 + index * 2
        sheet.merge_cells(start_row=3, start_column=column, end_row=3, end_column=column + 1)
        sheet.merge_cells(start_row=4, start_column=column, end_row=4, end_column=column + 1)
        label_cell = sheet.cell(3, column, safe_spreadsheet_text(label))
        value_cell = sheet.cell(4, column, safe_spreadsheet_text(value))
        label_cell.fill = HEADER_FILL
        label_cell.font = HEADER_FONT
        value_cell.fill = CARD_FILL
        value_cell.font = Font(bold=True, size=13)
        for cell in (label_cell, value_cell):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Periodo", "Código", "Asignatura", "Tipo", "Créditos", "Calificación"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(7, column, header)
    _style_header(sheet, 7, 1, len(headers))
    for course in semester.courses:
        sheet.append(
            [
                _display(course.period),
                _display(course.code),
                _display(course.name),
                _display(course.course_type),
                _display(course.credits),
                _display(course.grade),
            ]
        )
    last_course_row = 7 + len(semester.courses)
    _style_data_range(sheet, 8, last_course_row, len(headers))
    sheet.auto_filter.ref = f"A7:F{last_course_row}"
    sheet.freeze_panes = "A8"
    sheet.conditional_formatting.add(
        f"F8:F{last_course_row}",
        ColorScaleRule(  # type: ignore[no-untyped-call]
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=3,
            mid_color="FFEB84",
            end_type="num",
            end_value=5,
            end_color="63BE7B",
        ),
    )

    grade_chart = BarChart()
    grade_chart.type = "bar"
    grade_chart.title = "Calificaciones por asignatura"
    grade_chart.x_axis.title = "Calificación"
    grade_chart.add_data(
        Reference(sheet, min_col=6, min_row=7, max_row=last_course_row),
        titles_from_data=True,
    )
    grade_chart.set_categories(Reference(sheet, min_col=3, min_row=8, max_row=last_course_row))
    grade_chart.height = 8
    grade_chart.width = 15
    sheet.add_chart(grade_chart, "H7")

    type_row = max(last_course_row + 3, 24)
    sheet.cell(type_row, 8, "Tipo")
    sheet.cell(type_row, 9, "Créditos")
    _style_header(sheet, type_row, 8, 9)
    for offset, (course_type, credits) in enumerate(semester.credits_by_type, start=1):
        sheet.cell(type_row + offset, 8, safe_spreadsheet_text(course_type))
        sheet.cell(type_row + offset, 9, credits)
    if semester.credits_by_type:
        pie = PieChart()
        pie.title = "Créditos por tipo"
        pie.add_data(
            Reference(
                sheet, min_col=9, min_row=type_row, max_row=type_row + len(semester.credits_by_type)
            ),
            titles_from_data=True,
        )
        pie.set_categories(
            Reference(
                sheet,
                min_col=8,
                min_row=type_row + 1,
                max_row=type_row + len(semester.credits_by_type),
            )
        )
        pie.height = 7
        pie.width = 10
        sheet.add_chart(pie, f"K{type_row}")

    for column_letter, width in {
        "A": 30,
        "B": 16,
        "C": 42,
        "D": 22,
        "E": 12,
        "F": 14,
        "H": 24,
        "I": 12,
    }.items():
        sheet.column_dimensions[column_letter].width = width


def _write_curriculum_progress(workbook: Workbook, record: AcademicRecord) -> None:
    sheet = workbook.create_sheet("Avance curricular")
    sheet.sheet_properties.tabColor = ORANGE
    _style_title(sheet, "Avance por tipología", "H")
    headers = [
        "Tipología",
        "Exigidos",
        "Aprobados",
        "Pendientes",
        "Inscritos",
        "Cursados",
        "Avance",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, header)
    _style_header(sheet, 3, 1, len(headers))
    for item in record.credit_summaries:
        required = item.required.value if isinstance(item.required.value, float) else None
        passed = item.passed.value if isinstance(item.passed.value, float) else None
        progress = round(passed / required, 4) if required and passed is not None else None
        sheet.append(
            [
                _display(item.typology),
                _display(item.required),
                _display(item.passed),
                _display(item.pending),
                _display(item.enrolled),
                _display(item.taken),
                _metric(progress),
            ]
        )
    last_row = 3 + len(record.credit_summaries)
    _style_data_range(sheet, 4, last_row, len(headers))
    for cell in sheet["G"][3:]:
        if isinstance(cell.value, float):
            cell.number_format = "0.0%"
    if record.credit_summaries:
        chart = BarChart()
        chart.title = "Créditos aprobados y pendientes"
        chart.add_data(
            Reference(sheet, min_col=3, max_col=4, min_row=3, max_row=last_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=last_row))
        chart.height = 8
        chart.width = 14
        sheet.add_chart(chart, "I3")
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:G{last_row}"
    for column in range(1, 8):
        sheet.column_dimensions[get_column_letter(column)].width = 18


def _write_full_history(workbook: Workbook, record: AcademicRecord) -> None:
    sheet = workbook.create_sheet("Historial completo")
    sheet.sheet_properties.tabColor = BLUE
    _style_title(sheet, "Historial académico completo", "F")
    headers = ["Periodo", "Código", "Asignatura", "Tipo", "Créditos", "Calificación"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, header)
    _style_header(sheet, 3, 1, len(headers))
    for course in record.courses:
        sheet.append(
            [
                _display(course.period),
                _display(course.code),
                _display(course.name),
                _display(course.course_type),
                _display(course.credits),
                _display(course.grade),
            ]
        )
    last_row = 3 + len(record.courses)
    _style_data_range(sheet, 4, last_row, len(headers))
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:F{last_row}"
    for column_letter, width in {
        "A": 30,
        "B": 16,
        "C": 44,
        "D": 22,
        "E": 12,
        "F": 14,
    }.items():
        sheet.column_dimensions[column_letter].width = width


def _build_workbook(record: AcademicRecord) -> Workbook:
    workbook = Workbook()
    semesters = calculate_semester_metrics(record.courses)
    _write_summary(workbook, record, semesters)
    for semester in semesters:
        _write_semester(workbook, semester)
    _write_curriculum_progress(workbook, record)
    _write_full_history(workbook, record)
    return workbook


def export_xlsx(record: AcademicRecord, path: str | Path) -> Path:
    report = validate(record)
    if not report.valid:
        details = "; ".join(
            f"{issue.section or 'General'}: {issue.message}" for issue in report.issues
        )
        raise ValueError(f"No se exportó porque la extracción no está completa: {details}")

    destination = Path(path).expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=destination.parent,
            prefix=f".{destination.stem}-",
            suffix=".tmp.xlsx",
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        _build_workbook(record).save(temporary_path)
        secure_output_permissions(temporary_path)
        os.replace(temporary_path, destination)
        secure_output_permissions(destination)
        return destination
    except BaseException:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
        raise
