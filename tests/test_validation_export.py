from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sialytics.export import export_xlsx
from sialytics.models import (
    AcademicPeriod,
    AcademicRecord,
    Availability,
    CourseAttempt,
    Indicator,
    OfficialValue,
    Program,
    Provenance,
    SectionEvidence,
    Student,
)
from sialytics.validation import validate

SOURCE = "https://sia.unal.edu.co/ServiciosApp/faces/historia"
NOW = datetime(2026, 1, 2, tzinfo=UTC)


def value(label: str, raw: str | float) -> OfficialValue:
    return OfficialValue(
        Availability.AVAILABLE,
        Provenance(SOURCE, "Prueba", label, NOW),
        raw,
    )


def valid_record() -> AcademicRecord:
    program = Program("1", "Ingeniería de prueba", "Plan 1")
    course = CourseAttempt(
        period=value("Periodo académico", "2025-1S"),
        code=value("Código", "100001"),
        name=value("Asignatura", "=SUM(A1:A2)"),
        credits=value("Créditos", 3.0),
        grade=value("Calificación", 4.5),
        course_type=value("Tipo", "Disciplinar"),
    )
    period = AcademicPeriod(
        "2025-1S",
        {"Créditos registrados (Calculado por SIAlytics)": value("Calculado", 3.0)},
    )
    indicator = Indicator(
        "Promedio Aritmético Ponderado Acumulado",
        "PAPA",
        value("PAPA", 4.5),
    )
    evidence = {
        section: SectionEvidence(
            section,
            f"test-{index}",
            SOURCE,
            True,
            ("Campo",),
            1,
            True,
            1,
        )
        for index, section in enumerate(
            (
                "Estudiante",
                "Asignaturas",
                "Periodos académicos",
                "Tipologías",
                "Indicadores",
            )
        )
    }
    return AcademicRecord(
        Student(program, {"Nombre": value("Nombre", "Persona de prueba")}),
        [course],
        [period],
        [indicator],
        evidence,
        NOW,
    )


def test_valid_record_exports_atomic_workbook(tmp_path: Path) -> None:
    record = valid_record()
    destination = export_xlsx(record, tmp_path / "resultado.xlsx")
    workbook = load_workbook(destination)
    assert workbook.sheetnames == [
        "Resumen",
        "2025-1S",
        "Avance curricular",
        "Historial completo",
    ]
    assert workbook["2025-1S"]["C8"].value == "'=SUM(A1:A2)"
    assert workbook["2025-1S"]["A1"].value == "Semestre académico 2025-1S"
    assert workbook["2025-1S"]["A4"].value == 4.5
    assert workbook["2025-1S"]["E4"].value == 3
    assert workbook["Resumen"]["A3"].value == "PAPA"
    assert workbook["Resumen"]["A4"].value == 4.5
    assert workbook["Resumen"]["D3"].value == "PAPPI"
    assert workbook["Resumen"]["D4"].value == 4.5
    assert "Estado del dato" not in [cell.value for cell in workbook["2025-1S"][7]]
    assert "Fuente SIA" not in [cell.value for cell in workbook["2025-1S"][7]]
    student_facing_text = " ".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Calculado por SIAlytics" not in student_facing_text
    assert "Estado del dato" not in student_facing_text
    assert "Fuente SIA" not in student_facing_text
    assert "Nota" not in student_facing_text


def test_incomplete_record_is_not_exported(tmp_path: Path) -> None:
    record = valid_record()
    del record.evidence["Indicadores"]
    destination = tmp_path / "resultado.xlsx"
    with pytest.raises(ValueError, match="no está completa"):
        export_xlsx(record, destination)
    assert not destination.exists()
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_period_modalities_share_one_semester_sheet(tmp_path: Path) -> None:
    record = valid_record()
    record.courses.append(
        CourseAttempt(
            period=value("Periodo académico", "2025-1S Validación por suficiencia"),
            code=value("Código", "100002-Z"),
            name=value("Asignatura", "Prueba de suficiencia"),
            credits=value("Créditos", 4.0),
            grade=value("Calificación", 4.0),
            course_type=value("Tipo", "Disciplinar"),
        )
    )

    workbook = load_workbook(export_xlsx(record, tmp_path / "modalidades.xlsx"))

    assert "2025-1S" in workbook.sheetnames
    assert not any("suficiencia" in name.casefold() for name in workbook.sheetnames)
    assert workbook["2025-1S"]["G4"].value == 2
    assert workbook["2025-1S"]["A9"].value == "2025-1S Validación por suficiencia"


def test_ambiguous_empty_section_is_invalid() -> None:
    record = valid_record()
    record.evidence["Asignaturas"] = SectionEvidence(
        "Asignaturas", "test", SOURCE, True, ("Campo",), 1, True, 0
    )
    report = validate(record)
    assert not report.valid
    assert {issue.code for issue in report.issues} == {"ambiguous_empty"}


@pytest.mark.parametrize(
    "absence",
    [Availability.NOT_AVAILABLE, Availability.NO_RECORDED_INFORMATION],
)
def test_explicit_empty_section_is_valid(absence: Availability) -> None:
    record = valid_record()
    record.courses.clear()
    record.periods.clear()
    record.evidence["Asignaturas"] = SectionEvidence(
        "Asignaturas", "test", SOURCE, True, (), 1, True, 0, absence
    )
    record.evidence["Periodos académicos"] = SectionEvidence(
        "Periodos académicos", "test", SOURCE, True, (), 1, True, 0, absence
    )
    assert validate(record).valid


def test_official_section_absence_is_written_to_workbook(tmp_path: Path) -> None:
    record = valid_record()
    record.courses.clear()
    record.periods.clear()
    absence = Availability.NO_RECORDED_INFORMATION
    record.evidence["Asignaturas"] = SectionEvidence(
        "Asignaturas", "test", SOURCE, True, (), 1, True, 0, absence
    )
    record.evidence["Periodos académicos"] = SectionEvidence(
        "Periodos académicos", "test", SOURCE, True, (), 1, True, 0, absence
    )
    destination = export_xlsx(record, tmp_path / "ausencia.xlsx")
    workbook = load_workbook(destination)
    assert workbook.sheetnames == ["Resumen", "Avance curricular", "Historial completo"]
    assert workbook["Historial completo"].max_row == 3
